import tkinter as tk
from tkinter import messagebox, ttk
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

LANG = 'zh'
HISTORY_FILE = 'history.json'


def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_history(entries):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)


# === Excel generation (your original OpenPyXL UI) ===
def create_etf_interface(ws, user_inputs):
    ws['A1'] = "ETF智能估值分析系統"
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="4682B4")

    ws['A3'] = "📌 基本資料輸入"
    ws['A3'].font = Font(bold=True, size=12)

    input_labels = [
        "ETF代碼", "目前市價", "淨資產價值(NAV)", "市場趨勢"
    ]

    for i, label in enumerate(input_labels):
        ws.cell(row=4 + i, column=1, value=label)
        ws.cell(row=4 + i, column=2, value=user_inputs.get(label, ""))

    ws['A10'] = "✅ 產出時間"
    ws['B10'] = "=NOW()"

    # 下拉選單
    dv = DataValidation(type="list", formula1='"牛市,熊市,震盪,中性"', showDropDown=True)
    ws.add_data_validation(dv)
    ws["B7"].value = user_inputs.get("市場趨勢", "")
    dv.add(ws["B7"])


def create_etf_valuation_workbook(user_inputs, filename="ETF估值分析.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "ETF分析工具"
    create_etf_interface(ws, user_inputs)
    wb.save(filename)


# === GUI Labels ===
labels = {
    'zh': {
        "title": "ETF智能估值分析系統",
        "load": "載入",
        "save": "儲存",
        "generate": "產生Excel",
        "language": "English",
    },
    'en': {
        "title": "ETF Valuation Analysis System",
        "load": "Load",
        "save": "Save",
        "generate": "Generate Excel",
        "language": "中文",
    }
}


class ETFApp:
    def __init__(self, root):
        self.root = root
        self.root.title(labels[LANG]['title'])
        self.history_data = load_history()
        self.entry_vars = {}
        self.create_widgets()

    def create_widgets(self):
        global LANG
        self.fields = ["ETF代碼", "目前市價", "淨資產價值(NAV)", "市場趨勢"]

        for widget in self.root.winfo_children():
            widget.destroy()

        tk.Label(self.root, text=labels[LANG]['title'], font=("Arial", 16, "bold")).pack(pady=10)

        self.inputs_frame = tk.Frame(self.root)
        self.inputs_frame.pack()

        for field in self.fields:
            frame = tk.Frame(self.inputs_frame)
            frame.pack(fill='x', pady=5)

            lbl = tk.Label(frame, text=labels[LANG].get(field, field), width=20, anchor='w')
            lbl.pack(side='left', padx=5)

            var = tk.StringVar()
            ent = tk.Entry(frame, textvariable=var, width=30)
            ent.pack(side='left', padx=5, expand=True, fill='x')
            self.entry_vars[field] = var

        # History Dropdown
        self.history_var = tk.StringVar()
        self.history_combo = ttk.Combobox(self.root, textvariable=self.history_var, state='readonly')
        self.refresh_history_combo()
        self.history_combo.pack(pady=5)

        load_btn = tk.Button(self.root, text=labels[LANG]['load'], command=self.load_selected_history)
        load_btn.pack()

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        save_btn = tk.Button(btn_frame, text=labels[LANG]['save'], command=self.save_current_input)
        save_btn.pack(side='left', padx=5)

        generate_btn = tk.Button(btn_frame, text=labels[LANG]['generate'], command=self.generate_excel)
        generate_btn.pack(side='left', padx=5)

        lang_btn = tk.Button(self.root, text=labels[LANG]['language'], command=self.toggle_language)
        lang_btn.pack(pady=5)

    def toggle_language(self):
        global LANG
        LANG = 'en' if LANG == 'zh' else 'zh'
        self.create_widgets()

    def refresh_history_combo(self):
        display_list = [item.get("ETF代碼", "") for item in self.history_data]
        self.history_combo['values'] = display_list

    def load_selected_history(self):
        idx = self.history_combo.current()
        if idx < 0 or idx >= len(self.history_data):
            messagebox.showwarning("⚠️", "請選擇一筆資料" if LANG == 'zh' else "Please select an entry")
            return
        data = self.history_data[idx]
        for field in self.fields:
            self.entry_vars[field].set(data.get(field, ""))

    def save_current_input(self):
        entry = {}
        for field in self.fields:
            val = self.entry_vars[field].get().strip()
            if not val:
                messagebox.showerror("❌", f"{labels[LANG].get(field, field)} 不能為空" if LANG == 'zh' else f"{labels[LANG].get(field, field)} cannot be empty!")
                return
            entry[field] = val
        self.history_data.append(entry)
        save_history(self.history_data)
        self.refresh_history_combo()
        messagebox.showinfo("✅", "儲存成功" if LANG == 'zh' else "Saved successfully!")

    def generate_excel(self):
        try:
            # Defensive validation
            float(self.entry_vars["目前市價"].get())
            float(self.entry_vars["淨資產價值(NAV)"].get())

            user_inputs = {field: self.entry_vars[field].get() for field in self.fields}
            create_etf_valuation_workbook(user_inputs)
            messagebox.showinfo("✅", "Excel 產生成功" if LANG == 'zh' else "Excel generated successfully!")
        except ValueError:
            messagebox.showerror("❌", "請確認價格欄位為數字" if LANG == 'zh' else "Please ensure prices are valid numbers!")


if __name__ == "__main__":
    root = tk.Tk()
    app = ETFApp(root)
    root.mainloop()
